"""
Export orders to XLSX format.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from src.core.orders.models import Order, DeliveryType

logger = logging.getLogger(__name__)


class OrderExporter:
    """Export orders to XLSX format."""
    
    # Styles
    HEADER_FONT = Font(bold=True, size=14)
    SUBHEADER_FONT = Font(bold=True, size=11)
    NORMAL_FONT = Font(size=11)
    
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
    
    ALT_ROW_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
    LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
    RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
    WRAP_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    def export(self, order: Order, output_dir: Optional[Path] = None) -> Path:
        """
        Export order to XLSX file.
        
        Args:
            order: Order to export
            output_dir: Directory for output file (default: data/orders/)
            
        Returns:
            Path to created XLSX file
        """
        if output_dir is None:
            output_dir = Path("data/orders")
        
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Create filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"order_{order.id}_{timestamp}.xlsx"
        filepath = output_dir / filename
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Заявка {order.order_number}"
        
        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        
        row = 1
        
        # === HEADER ===
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws.cell(row=row, column=1, value=f"ЗАЯВКА {order.order_number}")
        cell.font = self.HEADER_FONT
        cell.alignment = self.CENTER_ALIGN
        row += 1
        
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws.cell(row=row, column=1, value=f"от {order.created_at.strftime('%d.%m.%Y %H:%M')}")
        cell.alignment = self.CENTER_ALIGN
        row += 2
        
        # === ITEMS TABLE ===
        ws.cell(row=row, column=1, value="ТОВАРЫ:").font = self.SUBHEADER_FONT
        row += 1
        
        # Table header
        headers = ["№", "Наименование", "Кол-во (кг)", "Цена/кг", "Сумма"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.HEADER_FONT_WHITE
            cell.fill = self.HEADER_FILL
            cell.border = self.THIN_BORDER
            cell.alignment = self.CENTER_ALIGN
        row += 1
        
        # Table data
        for i, item in enumerate(order.items, 1):
            # Item name with details
            name_parts = [item.product_name]
            if item.origin_country:
                name_parts.append(f"({item.origin_country})")
            name_parts.append(f"[{item.product_form}]")
            name = " ".join(name_parts)
            
            values = [
                i,
                name,
                f"{item.quantity_kg:.1f}",
                f"{item.price_per_kg:.0f} ₽",
                f"{item.total_price:.0f} ₽"
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.THIN_BORDER
                if col == 1:
                    cell.alignment = self.CENTER_ALIGN
                elif col in [3, 4, 5]:
                    cell.alignment = self.RIGHT_ALIGN
                else:
                    cell.alignment = self.LEFT_ALIGN
                
                # Alternate row coloring
                if i % 2 == 0:
                    cell.fill = self.ALT_ROW_FILL
            row += 1
        
        # Total row
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws.cell(row=row, column=1, value="ИТОГО:")
        cell.font = self.SUBHEADER_FONT
        cell.border = self.THIN_BORDER
        cell.alignment = self.RIGHT_ALIGN
        
        cell = ws.cell(row=row, column=3, value=f"{order.total_quantity:.1f}")
        cell.font = self.SUBHEADER_FONT
        cell.border = self.THIN_BORDER
        cell.alignment = self.RIGHT_ALIGN
        
        cell = ws.cell(row=row, column=4, value="")
        cell.border = self.THIN_BORDER
        
        cell = ws.cell(row=row, column=5, value=f"{order.total_price:.0f} ₽")
        cell.font = self.SUBHEADER_FONT
        cell.border = self.THIN_BORDER
        cell.alignment = self.RIGHT_ALIGN
        
        row += 2
        
        # === DELIVERY INFO ===
        if order.delivery:
            ws.cell(row=row, column=1, value="ДОСТАВКА:").font = self.SUBHEADER_FONT
            row += 1
            
            if order.delivery.delivery_type == DeliveryType.PICKUP:
                ws.merge_cells(f'B{row}:E{row}')
                ws.cell(row=row, column=1, value="Тип:")
                ws.cell(row=row, column=2, value="Самовывоз")
                row += 1
                ws.merge_cells(f'B{row}:E{row}')
                ws.cell(row=row, column=1, value="Адрес склада:")
                ws.cell(row=row, column=2, value="г. Киров, пер. Энгельса, 2")
            else:
                ws.merge_cells(f'B{row}:E{row}')
                ws.cell(row=row, column=1, value="Тип:")
                ws.cell(row=row, column=2, value="Доставка")
                row += 1
                
                if order.delivery.address:
                    ws.merge_cells(f'B{row}:E{row}')
                    ws.cell(row=row, column=1, value="Адрес:")
                    cell = ws.cell(row=row, column=2, value=order.delivery.address)
                    cell.alignment = self.WRAP_ALIGN
            row += 1
            
            if order.delivery.desired_date:
                weekday = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"][order.delivery.desired_date.weekday()]
                ws.merge_cells(f'B{row}:E{row}')
                ws.cell(row=row, column=1, value="Дата:")
                ws.cell(row=row, column=2, value=f"{order.delivery.desired_date.strftime('%d.%m.%Y')} ({weekday})")
                row += 1
                
                ws.merge_cells(f'B{row}:E{row}')
                ws.cell(row=row, column=1, value="Время:")
                ws.cell(row=row, column=2, value=order.delivery.format_time_slot())
            row += 2
        
        # === CUSTOMER INFO ===
        if order.customer:
            ws.cell(row=row, column=1, value="ЗАКАЗЧИК:").font = self.SUBHEADER_FONT
            row += 1
            
            if order.customer.name:
                ws.merge_cells(f'B{row}:E{row}')
                ws.cell(row=row, column=1, value="Имя:")
                ws.cell(row=row, column=2, value=order.customer.name)
                row += 1
            
            if order.customer.phone:
                ws.merge_cells(f'B{row}:E{row}')
                ws.cell(row=row, column=1, value="Телефон:")
                ws.cell(row=row, column=2, value=order.customer.phone)
                row += 1
            
            if order.customer.telegram_username:
                ws.merge_cells(f'B{row}:E{row}')
                ws.cell(row=row, column=1, value="Telegram:")
                ws.cell(row=row, column=2, value=f"@{order.customer.telegram_username}")
                row += 1
            
            if order.customer.company:
                ws.merge_cells(f'B{row}:E{row}')
                ws.cell(row=row, column=1, value="Компания:")
                ws.cell(row=row, column=2, value=order.customer.company)
            row += 2
        
        # === COMMENT ===
        if order.comment:
            ws.cell(row=row, column=1, value="КОММЕНТАРИЙ:").font = self.SUBHEADER_FONT
            row += 1
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws.cell(row=row, column=1, value=order.comment)
            cell.alignment = self.WRAP_ALIGN
            ws.row_dimensions[row].height = 40
        
        # Save
        wb.save(filepath)
        logger.info(f"Order exported to {filepath}")
        
        return filepath


# Singleton instance
order_exporter = OrderExporter()
